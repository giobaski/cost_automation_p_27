#!/usr/bin/env python
# -*- coding: utf-8 -*-

import arcpy, datetime, os, json, io
from datetime import datetime
from openpyxl import Workbook


arcpy.env.workspace = r'D:\Arcpy_pro_scripting\silk_projects\cost automation\საპროექტო_ფორმა_14\საპროექტო_ფორმა_13.gdb\Data'
arcpy.env.overwriteOutput = True


# STEP 1: გლობალური ცვლადების გამოცხადება
cable = 'Kabeli'
mowyobiloba = 'Mowkobiloba'
obieqti = 'obieqti'
xazi = 'Xazi'


CRM_ID = arcpy.GetParameterAsText(0)
crm_id = '777'


# STEP Create Layers
arcpy.MakeFeatureLayer_management(cable,'cable_layer', """ "CRM_ID" = {} OR "CRM_ID" is Null """.format(crm_id))
arcpy.MakeFeatureLayer_management(mowyobiloba,'mowyobiloba_layer',)
arcpy.MakeFeatureLayer_management(obieqti,'obieqti_layer',)
arcpy.MakeFeatureLayer_management(xazi,'xazi_layer',)



# result = arcpy.GetCount_management('cable_layer')
# print ("{} layer has {} records".format('cable_layer', result))
#
# with arcpy.da.SearchCursor('cable_layer', fields) as cable_cursor1:
#     for row in cable_cursor1:
#         print(row[2])


# STEP 2: კურსორისთვის საჭირო ველების ჩამონათვალი
fields = ['Type','SHAPE@LENGTH', 'CRM_ID']


# STEP 3:  "unique_domain_values"-ში ჩაიყრება უნიკალური დომეინის ტიპები sample: [40, 9, 11]
unique_domain_values = []
with arcpy.da.SearchCursor('cable_layer', fields) as cable_cursor:
    for row in cable_cursor:
        if row[0] not in unique_domain_values:
            unique_domain_values.append(row[0])




# STEP 4:  "length_by_domain_type" აქ ჩაიყრება კაბელის ტიპი და საერთო სიგრძე, sample: {40: 100.0, 9: 1500.0000121322373, 11: 50.00001876480431}
length_by_domain_type = {}
for i in unique_domain_values:
    length = 0
    with arcpy.da.SearchCursor('cable_layer', fields) as cable_cursor:
        for row in cable_cursor:
            if row[0] == i:
                length = length + row[1]
                length_by_domain_type[i] = round(length,1)
print(unique_domain_values)
print(length_by_domain_type)


# STEP 5: JSON DATA
with io.open("data/data.json", "r", encoding = "utf-8") as pricelist_file:
    data = json.loads(pricelist_file.read())
print("\n")




# STEP 6: ლუპი რომელიც სიგრძეებს გაამრავლებს data.json-ის ფაილში ჩაწერილ ფასებზე და შექმნის ახალ ლისტს ან დიქშენერის
cable_price_by_domain = {}
for key, value in length_by_domain_type.items():
    a = length_by_domain_type[key] * data[str(key)]["price"]
    cable_price_by_domain[key] = a
print("cable_price by domain: ")
print(cable_price_by_domain)


# STEP 7:
cable_price_by_name = {}
for key, value in cable_price_by_domain.items():
    newname = data[str(key)]["name"]
    cable_price_by_name[newname] = cable_price_by_domain[key]
print("cable price by name: ")
print(cable_price_by_name)





########################################################################################################################
# documentation https://openpyxl.readthedocs.io/en/stable/index.html
wb = Workbook()

# grab the active worksheet
ws = wb.active

ws.title = 'Cable'
ws.sheet_properties.tabColor = "1072BA"

# Data can be assigned directly to cells
ws['A1'] = 'ID'
ws['B1'] = 'კაბელის ტიპი'
ws['C1'] = 'ფასი'

# Rows can also be appended
id = 1
for key, value in cable_price_by_name.items():
    ws.append([id, key, value])
    id+=1

# Python types will automatically be converted
import datetime
ws['D2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
########################################################################################################################































# print 'CRM_ID XXX-ის ფარგლებში დახაზული კაბელების საერთო სიგრძეა: ' + str(total_length) + 'მეტრი, რომლის ღირებულებაც იქნება ' + str(cable_cost)
# result = arcpy.GetCount_management(cab_layer)
# print "{} layer has {} records".format(cab_layer,result)



#################################################### document ##########################################################

# # STEP 8: Create r/w output file
# outFile = io.open("outputs\Technical_Description.txt", "w")
# # Report header
# outFile.write("ტექპირობა")
# outFile.write(" \n")
# outFile.write("------------------------------------------------------------------------------- \n")
# outFile.write(" \n")
# # outFile.write('CRM_ID XXX-ის ფარგლებში დახაზული კაბელების საერთო სიგრძეა: ' + str(total_length) + 'მეტრი, რომლის ღირებულებაც იქნება ' + str(cable_cost))
# for key, values in cable_price_by_name.items():
#     outFile.write("კაბელის ტიპი:" + "'" + str(key) + "'" + ", ღირებულება: " + str(values) + " ლარი" + "\n")
# outFile.write(" \n")
# outFile.write("------------------------------------------------------------------------------- \n")
# outFile.write("თარიღი: " + str(datetime.datetime.today().strftime("%B %d, %Y")) + "\n")
#
# outFile.close()
# os.startfile("outputs\Technical_Description.txt")

################################# END DOCUMENT #########################################################################





